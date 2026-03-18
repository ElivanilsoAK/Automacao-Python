import os
import time
import datetime
import shutil
import zipfile
import threading
import schedule
import pandas as pd
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from selenium import webdriver
import sys
from PIL import Image, ImageDraw, ImageFont
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import ElementClickInterceptedException, TimeoutException, NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager
import io

class AutomationWorker(threading.Thread):
    def __init__(self, config, logger_callback=None, status_callback=None, clear_log_callback=None, update_stats_callback=None):
        super().__init__()
        self.config = config
        self.log_callback = logger_callback
        self.status_callback = status_callback
        self.clear_log_callback = clear_log_callback
        self.update_stats_callback = update_stats_callback
        self.success_count = 0
        self.fail_count = 0
        self.stop_flag = False
        self.daemon = True

    def log(self, message):
        if self.log_callback:
            self.log_callback(message)
        else:
            print(message)

    def status(self, message):
        if self.status_callback:
            self.status_callback(message)

    def setup_driver(self):
        chrome_options = Options()

        # ── Flags de segurança/SSL (sempre ativas) ──────────────────────────
        chrome_options.add_argument("--ignore-certificate-errors")
        chrome_options.add_argument("--ignore-ssl-errors")
        chrome_options.add_argument("--allow-insecure-localhost")

        # ── Flags críticas de estabilidade (SEMPRE, não só no headless) ─────
        # Sem estas flags o Chrome encerra imediatamente quando iniciado via
        # Selenium a partir de um EXE ou sessão não interativa do Windows.
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--disable-extensions")
        chrome_options.add_argument("--disable-infobars")
        chrome_options.add_argument("--disable-popup-blocking")
        chrome_options.add_argument("--remote-debugging-port=0")   # porta aleatória livre
        chrome_options.add_argument("--log-level=3")               # suprime logs desnecessários

        if self.config.get("headless", False):
            chrome_options.add_argument("--headless=new")
            chrome_options.add_argument("--disable-software-rasterizer")
            chrome_options.add_argument("--window-size=1920,1080")
        else:
            chrome_options.add_argument("--start-maximized")

        # ── Chrome binary: tenta caminhos padrão se não encontrar ───────────
        chrome_paths = [
            r"C:\Program Files\Google\Chrome\Application\chrome.exe",
            r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
            r"C:\Users\elivanilso.junior\AppData\Local\Google\Chrome\Application\chrome.exe",
        ]
        for p in chrome_paths:
            if os.path.exists(p):
                chrome_options.binary_location = p
                self.log(f"Chrome encontrado em: {p}")
                break

        # ── Pasta de download ────────────────────────────────────────────────
        raw_download_dir = self.config.get("download_path", os.getcwd())
        download_dir = os.path.abspath(raw_download_dir)

        if not os.path.exists(download_dir):
            try:
                os.makedirs(download_dir)
                self.log(f"Pasta criada: {download_dir}")
            except Exception as e:
                self.log(f"Erro ao criar pasta: {e}")

        prefs = {
            "download.default_directory": download_dir,
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "safebrowsing.enabled": False,
            "profile.default_content_settings.popups": 0,
        }
        chrome_options.add_experimental_option("prefs", prefs)

        # ── Inicia o driver ──────────────────────────────────────────────────
        try:
            driver = webdriver.Chrome(
                service=Service(ChromeDriverManager().install()),
                options=chrome_options
            )
        except Exception as e:
            self.log(f"Erro ao criar driver: {e}")
            raise

        return driver, download_dir

    def safe_click(self, driver, element):
        """Tenta clique normal, se falhar usa JS."""
        try:
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
            time.sleep(0.3)
            element.click()
        except Exception as e:
            self.log(f"Clique normal falhou ({str(e)[:50]}...), tentando JS...")
            driver.execute_script("arguments[0].click();", element)

    def _close_modal_if_present(self, driver):
        """
        FIX: Fecha o modal 'Primeiros Passos' que aparece após o login
        e bloqueia interações com a página.
        """
        try:
            # Aguarda até 5 segundos se o modal aparecer
            modal_close = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, ".ui-dialog-titlebar-close, .p-dialog-header-close, button[aria-label='Close'], .ui-dialog .ui-dialog-titlebar button"))
            )
            self.safe_click(driver, modal_close)
            self.log("Modal 'Primeiros Passos' fechado.")
            time.sleep(1)
        except TimeoutException:
            pass  # Nenhum modal presente — ok
        except Exception as e:
            self.log(f"Aviso ao fechar modal: {str(e)[:80]}")

    def _login(self, driver, wait):
        self.status("Realizando login...")
        try:
            # FIX: Usar IDs exatos do site InControl em vez de seletores genéricos
            # IDs confirmados via inspeção do DOM: input_login_usuario, input_login_senha, btn_login
            try:
                user_input = wait.until(EC.presence_of_element_located((By.ID, "input_login_usuario")))
            except TimeoutException:
                # Fallback para seletor genérico caso o sistema mude
                user_input = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='text'], input[name='username']")))

            if not user_input.is_displayed():
                self.log("Já logado ou tela de login não encontrada.")
                return

            self.log("Inserindo credenciais...")
            user_input.clear()
            user_input.send_keys(self.config["incontrol_user"])

            try:
                pass_input = driver.find_element(By.ID, "input_login_senha")
            except NoSuchElementException:
                pass_input = driver.find_element(By.CSS_SELECTOR, "input[type='password']")

            pass_input.clear()
            pass_input.send_keys(self.config["incontrol_password"])

            try:
                login_btn = driver.find_element(By.ID, "btn_login")
            except NoSuchElementException:
                login_btn = driver.find_element(By.CSS_SELECTOR, "button[type='submit']")

            self.safe_click(driver, login_btn)
            self.log("Login submetido. Aguardando carregamento...")
            time.sleep(6)

            # FIX: Fechar modal de boas-vindas que bloqueia toda a interface
            self._close_modal_if_present(driver)

        except TimeoutException:
            self.log("Tela de login não encontrada (já logado ou URL diferente).")
        except Exception as e:
            self.log(f"Aviso no login: {e}")

    def _apply_date_filters(self, driver, wait):
        self.log("Aplicando filtro de data de HOJE...")
        try:
            # 0. Abrir Painel de Filtros
            try:
                btn_abrir = wait.until(EC.element_to_be_clickable((By.ID, "bnt_filtro_v3_filtros")))
                self.safe_click(driver, btn_abrir)
                time.sleep(1.5)
            except TimeoutException:
                self.log("Botão de filtros não encontrado. Pulando filtro de data.")
                return

            now = datetime.datetime.now()
            start_str = now.strftime("%d/%m/%Y 00:00")
            end_str = now.strftime("%d/%m/%Y %H:%M")

            def set_date(placeholder, value):
                inp = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, f"input[placeholder='{placeholder}']")))
                self.safe_click(driver, inp)
                # Limpa com múltiplas estratégias para garantir
                driver.execute_script("arguments[0].value = '';", inp)
                inp.send_keys(Keys.CONTROL + "a")
                inp.send_keys(Keys.DELETE)
                time.sleep(0.2)
                inp.send_keys(value)
                inp.send_keys(Keys.TAB)
                time.sleep(0.3)

            set_date("Data inicial dos eventos", start_str)
            set_date("Data final dos eventos", end_str)

            # Botão Filtrar
            btn_filtrar = wait.until(EC.element_to_be_clickable((By.ID, "bnt_filtro_v3_filtrar")))
            self.safe_click(driver, btn_filtrar)
            self.log(f"Filtro de data submetido: {start_str} até {end_str}")
            time.sleep(5)

        except Exception as e:
            self.log(f"Erro ao aplicar datas: {e}")

    def _trigger_export(self, driver, wait):
        self.log("Acionando menu de exportação...")
        try:
            # Estratégia 1: Seta do SplitButton (abre menu com opções CSV/PDF)
            dropdown_arrow = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[contains(@class, 'ui-splitbutton-menubutton')]")))
            self.safe_click(driver, dropdown_arrow)
            time.sleep(1)

            # Estratégia 2: Item CSV no menu dropdown
            csv_opt = wait.until(EC.element_to_be_clickable((By.XPATH, "//span[contains(text(), 'CSV')] | //a[contains(text(), 'CSV')] | //li[contains(., 'CSV')]")))
            self.safe_click(driver, csv_opt)
            self.log("Opção CSV selecionada.")
            return True
        except Exception as e:
            self.log(f"Falha no menu dropdown CSV: {e}. Tentando clique direto no botão principal...")
            try:
                # Fallback: clica no botão principal "Exportar relatório"
                export_btn = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[contains(.,'Exportar') or contains(.,'Exportar relatório')]")))
                self.safe_click(driver, export_btn)
                self.log("Botão exportar clicado (modo fallback).")
                return True
            except Exception as e2:
                self.log(f"Erro ao clicar exportar: {e2}")
                return False

    def login_and_export_report(self):
        self.log("Iniciando extração...")
        self.status("Extraindo dados...")

        driver = None
        current_download_dir = self.config.get("download_path", os.getcwd())
        if not os.path.exists(current_download_dir):
            os.makedirs(current_download_dir)

        try:
            driver, _ = self.setup_driver()
            driver.get(self.config["incontrol_url"])
            wait = WebDriverWait(driver, 30)

            self._login(driver, wait)

            # Garante que está na página correta
            if "eventos-usuario" not in driver.current_url:
                self.log("Navegando para página de eventos...")
                driver.get(self.config["incontrol_url"])
                time.sleep(5)
                self._close_modal_if_present(driver)

            # Aplica filtro de data
            self._apply_date_filters(driver, wait)

            # Marca tempo ANTES do clique de exportação
            start_export_time = time.time()

            if self._trigger_export(driver, wait):
                self.log("Aguardando download...")
                return self._wait_for_download(current_download_dir, start_export_time)

            return None

        except Exception as e:
            self.log(f"Erro no fluxo Selenium: {e}")
            return None
        finally:
            if driver:
                try:
                    driver.quit()
                except:
                    pass

    def _wait_for_download(self, download_dir, start_time, timeout=120):
        loop_start = time.time()
        self.log(f"Monitorando pasta: {download_dir}")
        while time.time() - loop_start < timeout:
            if self.stop_flag:
                return None
            try:
                files = os.listdir(download_dir)
                # .crdownload = download ainda em progresso
                if any(f.endswith('.crdownload') for f in files):
                    time.sleep(1)
                    continue

                # Candidatos: arquivos CSV modificados após início do export
                candidates = []
                for f in files:
                    full_path = os.path.join(download_dir, f)
                    if f.lower().endswith('.csv') and os.path.getmtime(full_path) > (start_time - 15):
                        candidates.append(full_path)

                if candidates:
                    latest = max(candidates, key=os.path.getmtime)
                    self.log(f"Download concluído: {os.path.basename(latest)}")
                    return self.organize_files(latest)

            except Exception:
                pass
            time.sleep(1)
        self.log("Timeout: download não detectado em 120 segundos.")
        return None

    def _get_report_dir(self):
        """
        Retorna a pasta do relatório do dia e o timestamp atual.

        Estrutura criada:
          <download_path>/
            relatorios/
              YYYY-MM-DD/          ← execuções do dia
            report-atual.csv       ← sempre o CSV mais recente
            backups/
              backup-YYYY-MM.zip   ← arquivo mensal automático
        """
        base = os.path.abspath(self.config.get("download_path", os.getcwd()))
        now = datetime.datetime.now()
        date_str = now.strftime("%Y-%m-%d")
        time_str = now.strftime("%H%M%S")

        report_dir = os.path.join(base, "relatorios", date_str)
        os.makedirs(report_dir, exist_ok=True)

        backup_dir = os.path.join(base, "backups")
        os.makedirs(backup_dir, exist_ok=True)

        return base, report_dir, time_str, now

    def organize_files(self, downloaded_file_path):
        """
        Move o CSV baixado para a pasta organizada do dia e atualiza report-atual.csv.
        Retorna um dict com todos os paths relevantes para esta execução.
        """
        base, report_dir, time_str, now = self._get_report_dir()

        # 1. Copiar CSV original → relatorios/YYYY-MM-DD/raw-HHMMSS.csv
        raw_name = f"raw-{time_str}.csv"
        raw_path = os.path.join(report_dir, raw_name)
        try:
            time.sleep(1)  # garante que o Chrome finalizou a escrita
            shutil.copy2(downloaded_file_path, raw_path)
            self.log(f"CSV raw salvo: relatorios/{now.strftime('%Y-%m-%d')}/{raw_name}")
        except Exception as e:
            self.log(f"Aviso ao copiar CSV raw: {e}")
            raw_path = downloaded_file_path

        # 2. Atualiza report-atual.csv na raiz (referência rápida)
        atual_path = os.path.join(base, "report-atual.csv")
        try:
            shutil.copy2(raw_path, atual_path)
        except Exception as e:
            self.log(f"Aviso ao atualizar report-atual.csv: {e}")

        # 3. Remove o arquivo temporário original do download
        try:
            if os.path.abspath(downloaded_file_path) != os.path.abspath(raw_path):
                os.remove(downloaded_file_path)
        except Exception:
            pass

        # 4. Backup mensal automático de CSVs do dia anterior ou mais antigos
        self.archive_old_reports(base, now)

        return {
            "atual": atual_path,
            "raw": raw_path,
            "report_dir": report_dir,
            "time_str": time_str,
            "date_str": now.strftime("%Y-%m-%d"),
        }

    def archive_old_reports(self, base_dir, now):
        """
        Compacta as pastas de relatórios de meses anteriores em ZIPs mensais
        dentro de backups/, mantendo o dia atual intocado.
        """
        relatorios_base = os.path.join(base_dir, "relatorios")
        if not os.path.exists(relatorios_base):
            return

        backup_dir = os.path.join(base_dir, "backups")
        current_month = now.strftime("%Y-%m")
        current_date = now.strftime("%Y-%m-%d")

        try:
            os.makedirs(backup_dir, exist_ok=True)
            for date_folder in os.listdir(relatorios_base):
                full_folder = os.path.join(relatorios_base, date_folder)
                if not os.path.isdir(full_folder):
                    continue

                # Data do passado (não hoje)
                folder_month = date_folder[:7]  # YYYY-MM
                if date_folder == current_date:
                    continue  # nunca arquiva o dia atual

                if folder_month < current_month:
                    # Arquiva no ZIP do mês correspondente
                    zip_name = f"backup-{folder_month}.zip"
                    zip_path = os.path.join(backup_dir, zip_name)
                    files_in = [f for f in os.listdir(full_folder)
                                if os.path.isfile(os.path.join(full_folder, f))]
                    if files_in:
                        self.log(f"Arquivando {date_folder} em {zip_name}...")
                        with zipfile.ZipFile(zip_path, 'a', zipfile.ZIP_DEFLATED) as zipf:
                            for fn in files_in:
                                src = os.path.join(full_folder, fn)
                                arcname = os.path.join(date_folder, fn)
                                zipf.write(src, arcname)
                        # Remove pasta após zipar
                        shutil.rmtree(full_folder, ignore_errors=True)
        except Exception as e:
            self.log(f"Aviso no backup mensal: {e}")

    def generate_graphs(self, df, df_inside):
        """Gera gráficos usando Matplotlib (import lazy)."""
        try:
            import matplotlib
            matplotlib.use('Agg')
            import matplotlib.pyplot as plt
        except ImportError:
            self.log("Matplotlib não instalado. Gráficos ignorados.")
            return {}

        images = {}
        try:
            plt.style.use('ggplot')

            if not df.empty and 'Data Evento' in df.columns:
                plt.figure(figsize=(8, 4))
                df['Hora'] = df['Data Evento'].dt.hour
                hourly = df.groupby('Hora').size()
                hourly.plot(kind='line', marker='o', color='#198754', linestyle='-', linewidth=2)
                plt.title('Fluxo de Acessos por Hora', fontsize=12, color='#333')
                plt.xlabel('Hora do Dia')
                plt.ylabel('Movimentos')
                plt.grid(True, linestyle='--', alpha=0.6)
                plt.tight_layout()

                buf_line = io.BytesIO()
                plt.savefig(buf_line, format='png')
                buf_line.seek(0)
                images['line_chart'] = buf_line.read()
                plt.close()

        except Exception as e:
            self.log(f"Erro ao gerar gráficos: {e}")

        return images

    def _map_columns(self, df):
        """
        FIX: Mapeamento de colunas com prioridade por nome exato do CSV do InControl,
        antes de recorrer a fuzzy matching. Evita capturar coluna errada.

        Colunas do CSV InControl confirmadas:
        Nome do Usuário | RG | CPF | Matrícula | Departamento | Tipo |
        Nome visitado | Código da Credencial | Acesso | Nome do Dispositivo |
        Ponto de Acesso | Data Evento | Tipo de evento | Status | Observação | ...
        """
        df.columns = df.columns.str.strip()
        
        # Mapeamento EXATO por nome de coluna (prioridade máxima)
        exact_map = {
            'Nome do Usuário': 'Nome',
            'Nome do Usuario': 'Nome',
            'Departamento': 'Departamento',
            'Acesso': 'Direcao',          # "Entrada" / "Saída"
            'Status': 'Status',            # "Acesso Liberado" / "Acesso Negado"
            'Tipo de evento': 'TipoEvento',# "Acesso Liberado" / "Acesso Negado"
            'Ponto de Acesso': 'Local',
            'Data Evento': 'Data',
            'CPF': 'CPF',
            'Matrícula': 'Matricula',
            'Matricula': 'Matricula',
        }
        
        rename_map = {}
        used_targets = set()
        
        # Passo 1: Match exato
        for col in df.columns:
            col_stripped = col.strip()
            if col_stripped in exact_map:
                target = exact_map[col_stripped]
                if target not in used_targets:
                    rename_map[col] = target
                    used_targets.add(target)
        
        # Passo 2: Fuzzy fallback para colunas ainda não mapeadas
        fuzzy_targets = {
            'Nome': ['nome do usuário', 'nome do usuario', 'nome', 'usuario', 'colaborador'],
            'Departamento': ['departamento', 'setor', 'area', 'empresa'],
            'Status': ['status', 'situacao'],
            'Direcao': ['acesso'],  # Apenas 'acesso' aqui — muito específico
            'Local': ['ponto de acesso', 'local', 'porta'],
            'Data': ['data evento', 'data do evento', 'data', 'horario'],
            'CPF': ['cpf', 'documento'],
            'Matricula': ['matricula', 'cracha', 'código da credencial'],
        }
        
        for target, candidates in fuzzy_targets.items():
            if target in used_targets:
                continue
            for col in df.columns:
                if col in rename_map:
                    continue
                col_lower = col.lower().strip()
                for cand in candidates:
                    if cand in col_lower:
                        rename_map[col] = target
                        used_targets.add(target)
                        break
                if target in used_targets:
                    break
        
        df = df.rename(columns=rename_map)
        self.log(f"Colunas mapeadas: {rename_map}")
        return df

    def process_data(self, paths_or_csv):
        """
        Aceita tanto o dict retornado por organize_files quanto uma string de path
        (modo compatível com chamadas diretas).
        Salva o CSV processado e a lista de presentes na pasta do dia.
        """
        self.status("Analisando Dados...")

        # Resolve path e pasta de destino
        if isinstance(paths_or_csv, dict):
            csv_path = paths_or_csv["atual"]
            report_dir = paths_or_csv.get("report_dir", os.path.dirname(csv_path))
            time_str = paths_or_csv.get("time_str", datetime.datetime.now().strftime("%H%M%S"))
        else:
            csv_path = paths_or_csv
            base, report_dir, time_str, _ = self._get_report_dir()

        try:
            # 1. Carregar CSV
            try:
                df = pd.read_csv(csv_path, sep=';', encoding='utf-8', on_bad_lines='skip')
            except UnicodeDecodeError:
                self.log("Falha UTF-8. Tentando Latin-1...")
                df = pd.read_csv(csv_path, sep=';', encoding='latin1', on_bad_lines='skip')

            self.log(f"CSV carregado: {len(df)} linhas, colunas: {list(df.columns)}")

            # 2. Mapeamento de Colunas (FIX)
            df = self._map_columns(df)

            # 3. Tratamento de colunas mínimas e vazias
            for col in ['Nome', 'Departamento', 'Status', 'Data', 'CPF', 'Matricula']:
                if col not in df.columns:
                    df[col] = ''

            # 4. Tratamento de Departamento vazio (Visitantes / Desconhecidos)
            if 'Departamento' in df.columns:
                df['Departamento'] = df['Departamento'].fillna('').astype(str).str.strip()
                # Substitui valores vazios ou inválidos por VISITANTE / NÃO INFORMADO
                df.loc[df['Departamento'].isin(['', 'nan', 'N/A', 'null', 'None']), 'Departamento'] = 'VISITANTE / NÃO INFORMADO'

            # 5. Filtro: Status "Acesso Liberado"
            if 'Status' in df.columns:
                initial_count = len(df)
                df = df.dropna(subset=['Status'])
                df = df[df['Status'].astype(str).str.contains("Liberado", case=False, na=False)]
                self.log(f"Filtro Status: {initial_count} → {len(df)} (apenas 'Acesso Liberado')")
            else:
                self.log("AVISO: Coluna 'Status' não encontrada. Filtragem não aplicada.")

            # 6. Filtro: Remover usuários inválidos (apenas nomes claramente errados)
            if 'Nome' in df.columns:
                df = df[~df['Nome'].astype(str).str.lower().str.contains(
                    r'^desconhecido$|^n/a$|^usuario desconhecido$', na=False, regex=True
                )]

            # 7. Selecionar apenas colunas normalizadas
            keep_cols = [c for c in ['Nome', 'Departamento', 'Status', 'Direcao', 'Local', 'Data', 'CPF', 'Matricula'] if c in df.columns]
            df = df[keep_cols].copy()
            df = df.loc[:, ~df.columns.duplicated()]

            # 7. Conversão de Data
            df['Data Evento'] = pd.to_datetime(df['Data'], errors='coerce')
            df = df.dropna(subset=['Data Evento'])

            # 8. Filtro Hoje
            hoje = datetime.datetime.now().date()
            df = df[df['Data Evento'].dt.date == hoje]
            self.log(f"Após filtro de hoje ({hoje}): {len(df)} registros")

            if df.empty:
                self.log("AVISO: Nenhum registro para hoje no CSV. Verifique o filtro de data no site.")

            # 9. FIX CRÍTICO — Lógica de Presença com identidade robusta
            # Problema original: groupby com CPF/Matrícula NaN cria chaves duplicadas inválidas.
            # Normalizar NaN → string vazia ANTES do groupby, usar Nome + CPF_norm + Mat_norm.
            df['CPF_norm'] = df['CPF'].fillna('').astype(str).str.strip()
            df['Mat_norm'] = df['Matricula'].fillna('').astype(str).str.strip()
            df['Nome_norm'] = df['Nome'].astype(str).str.strip().str.upper()

            def make_identity(row):
                nome = row['Nome_norm']
                cpf = row['CPF_norm']
                mat = row['Mat_norm']
                if cpf and cpf not in ('', 'N/A', 'nan'):
                    return f"{nome}|{cpf}"
                elif mat and mat not in ('', 'N/A', 'nan', '0'):
                    return f"{nome}|MAT:{mat}"
                else:
                    return nome

            df['identity_key'] = df.apply(make_identity, axis=1)

            # 10. Deduplicação Temporal (eventos < 60s da mesma pessoa logados duplicados)
            # Ordenar pela chave de identidade e tempo
            df_sorted = df.sort_values(by=['identity_key', 'Data Evento'])
            df_sorted['TimeDiff'] = df_sorted.groupby('identity_key')['Data Evento'].diff()
            df_sorted = df_sorted[(df_sorted['TimeDiff'].isnull()) | (df_sorted['TimeDiff'] > pd.Timedelta(seconds=60))]

            # 11. Pega o ÚLTIMO evento registrado por pessoa
            df_last = df_sorted.groupby('identity_key', sort=False).last().reset_index()

            # 11. Quem está DENTRO? (Último evento = "Entrada")
            if 'Direcao' in df_last.columns:
                # Normaliza: remove acentos, lowercase
                df_last['Direcao_norm'] = (
                    df_last['Direcao'].astype(str)
                    .str.lower()
                    .str.normalize('NFKD')
                    .str.encode('ascii', errors='ignore')
                    .str.decode('utf-8')
                    .str.strip()
                )

                # "Entrada" → dentro; "Saída/Saida" → fora
                df_inside = df_last[
                    df_last['Direcao_norm'].str.contains(r'^entr', na=False, regex=True)
                ].copy()

                self.log(f"Presença calculada: {len(df_inside)} pessoa(s) dentro da obra.")
                self.log(f"  Exemplos: {list(df_inside['Nome'].head(5))}")
            else:
                self.log("ERRO: Coluna de direção (Acesso/Entada/Saída) não encontrada.")
                df_inside = pd.DataFrame()

            # 12. Estatísticas
            total_presentes = len(df_inside)
            total_movimentos = len(df)
            self.log(f"Total de movimentos: {total_movimentos} | Presentes: {total_presentes}")

            # --- SALVAR CSV PROCESSADO (auditoria) ---
            try:
                proc_path = os.path.join(report_dir, f"processado-{time_str}.csv")
                df_sorted.to_csv(proc_path, sep=';', encoding='utf-8-sig', index=False)
                self.log(f"CSV processado salvo: processado-{time_str}.csv")
            except Exception as e:
                self.log(f"Aviso ao salvar CSV processado: {e}")

            # --- GERAR PLANILHA EXCEL COM PRESENTES (para email e auditoria) ---
            xlsx_path = None
            try:
                xlsx_path = self._gerar_planilha_presentes(df_inside, report_dir, time_str)
            except Exception as e:
                self.log(f"Aviso ao gerar planilha Excel: {e}")

            # 13. Tabela por Departamento
            dept_col = 'Departamento'
            if dept_col in df_inside.columns:
                dept_stats = df_inside.groupby(dept_col).size().reset_index(name='Qtd')
                dept_stats = dept_stats.sort_values('Qtd', ascending=False)
                total_dept = dept_stats['Qtd'].sum()
                dept_stats['Pct'] = (dept_stats['Qtd'] / total_dept * 100).fillna(0).round(1)
            else:
                dept_stats = pd.DataFrame(columns=[dept_col, 'Qtd', 'Pct'])

            # 14. HTML da tabela
            rows_html = ""
            for _, row in dept_stats.iterrows():
                pct_val = row['Pct']
                rows_html += f"""
                <tr style="border-bottom: 1px solid #f0f0f0;">
                    <td style="padding: 14px 20px; font-size: 14px; color: #333;">{row[dept_col]}</td>
                    <td style="padding: 14px 20px; text-align: right; font-weight: bold; color: #333; font-size: 14px;">{row['Qtd']}</td>
                    <td style="padding: 14px 20px; text-align: right;">
                        <span style="background-color: #e8f5e9; color: #198754; padding: 4px 10px; border-radius: 12px; font-size: 12px; font-weight: 600;">{pct_val:.1f}%</span>
                    </td>
                </tr>
                """

            # 15. Gráficos
            graph_imgs = {}
            try:
                graph_imgs = self.generate_graphs(df, df_inside)
            except Exception as e:
                self.log(f"Erro ao gerar gráficos: {e}")

            graphs_html = ""
            if 'line_chart' in graph_imgs:
                graphs_html = """
                    <h3 style="color: #198754; font-size: 16px; margin-top: 30px; border-left: 4px solid #ffc107; padding-left: 10px;">Fluxo Horário</h3>
                    <div style="text-align: center; margin-top: 15px;">
                        <img src="cid:line_chart" style="max-width: 100%; border-radius: 8px; box-shadow: 0 2px 5px rgba(0,0,0,0.05);">
                    </div>
                """

            # 16. Data em português
            meses_pt = {
                1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril', 5: 'Maio', 6: 'Junho',
                7: 'Julho', 8: 'Agosto', 9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro'
            }
            hoje_dt = datetime.datetime.now()
            data_pt = f"{hoje_dt.day} de {meses_pt[hoje_dt.month]} de {hoje_dt.year}"

            # 17. HTML do Relatório
            html_body = f"""
            <html>
            <body style="font-family: 'Segoe UI', 'Roboto', Helvetica, Arial, sans-serif; background-color: #f4f6f8; margin: 0; padding: 20px;">

                <div style="max-width: 650px; margin: 0 auto; background-color: #ffffff; border-radius: 16px; box-shadow: 0 8px 30px rgba(0,0,0,0.08); overflow: hidden;">

                    <!-- Header -->
                    <div style="background: linear-gradient(135deg, #06402b 0%, #1e5c3f 100%); color: white; padding: 40px 30px; text-align: center;">
                        <h2 style="margin: 0; font-size: 26px; font-weight: 500; letter-spacing: 0.5px;">Relatório Diário de Obra</h2>
                        <div style="width: 50px; height: 3px; background-color: #ffc107; margin: 15px auto; border-radius: 2px;"></div>
                        <p style="margin: 10px 0 0 0; font-size: 16px; opacity: 0.9;">Monitoramento de Acesso</p>
                        <div style="margin-top: 20px; font-size: 14px; color: #e8f5e9; font-weight: 500;">
                            📅 {data_pt}
                        </div>
                    </div>

                    <div style="padding: 40px 30px;">

                        <!-- Cards -->
                        <div style="display: flex; gap: 20px; margin-bottom: 40px;">
                            <div style="flex: 1; padding: 25px 20px; border-radius: 12px; text-align: center; border: 1px solid #e0e0e0; background: #ffffff; box-shadow: 0 2px 8px rgba(0,0,0,0.03);">
                                <div style="color: #555; font-size: 13px; text-transform: uppercase; letter-spacing: 1px; font-weight: 600; margin-bottom: 8px;">Total de Movimentos</div>
                                <div style="font-size: 36px; font-weight: 700; color: #2c3e50;">{total_movimentos}</div>
                            </div>
                            <div style="flex: 1; padding: 25px 20px; border-radius: 12px; text-align: center; border: 1px solid #c3e6cb; background: #f1f8f3; box-shadow: 0 4px 12px rgba(25, 135, 84, 0.1);">
                                <div style="color: #155724; font-size: 13px; text-transform: uppercase; letter-spacing: 1px; font-weight: 600; margin-bottom: 8px;">Pessoas na Obra</div>
                                <div style="font-size: 36px; font-weight: 700; color: #198754;">{total_presentes}</div>
                            </div>
                        </div>

                        <!-- Tabela -->
                        <div style="margin-bottom: 40px;">
                            <h3 style="color: #2c3e50; font-size: 18px; margin-bottom: 20px; padding-left: 15px; border-left: 4px solid #198754; font-weight: 600;">
                                Presença por Departamento
                            </h3>
                            <table style="width: 100%; border-collapse: separate; border-spacing: 0;">
                                <thead>
                                    <tr style="background-color: #fafafa;">
                                        <th style="padding: 15px 20px; text-align: left; color: #7f8c8d; font-size: 12px; font-weight: 600; text-transform: uppercase; border-bottom: 2px solid #eee;">Departamento</th>
                                        <th style="padding: 15px 20px; text-align: right; color: #7f8c8d; font-size: 12px; font-weight: 600; text-transform: uppercase; border-bottom: 2px solid #eee;">Qtd.</th>
                                        <th style="padding: 15px 20px; text-align: right; color: #7f8c8d; font-size: 12px; font-weight: 600; text-transform: uppercase; border-bottom: 2px solid #eee;">%</th>
                                    </tr>
                                </thead>
                                <tbody>
                                    {rows_html}
                                </tbody>
                            </table>
                        </div>

                        <!-- Gráfico -->
                        {graphs_html}

                    </div>

                    <!-- Footer -->
                    <div style="background-color: #2c3e50; padding: 25px; text-align: center;">
                        <p style="margin: 0; font-size: 12px; color: #bdc3c7; line-height: 1.5;">
                            Relatório gerado automaticamente pelo sistema <b>BotEliva</b>.<br>
                            Este é um email informativo, por favor não responda.
                        </p>
                    </div>
                </div>
            </body>
            </html>
            """

            # 18. Gerar Imagem Resumo — salva na pasta do dia
            summary_img_path = None
            try:
                summary_img_path = self.generate_summary_card(
                    total_movimentos, total_presentes, dept_stats, data_pt,
                    report_dir=report_dir, time_str=time_str
                )
            except Exception as e:
                self.log(f"Erro ao gerar imagem resumo: {e}")

            return {'html': html_body, 'images': graph_imgs,
                    'summary_image': summary_img_path, 'xlsx_presentes': xlsx_path}

        except Exception as e:
            import traceback
            self.log(f"ERRO CRÍTICO no processamento: {e}")
            self.log(traceback.format_exc())
            return f"Erro processamento dados: {e}"

    def generate_summary_card(self, total_mov, total_pres, df_dept, data_str,
                              report_dir=None, time_str=None):
        """Gera um card PNG de resumo e salva na pasta do relatório do dia."""
        if report_dir is None:
            report_dir = self.config.get("download_path", os.getcwd())
        if time_str is None:
            time_str = datetime.datetime.now().strftime("%H%M%S")
        try:
            width = 800
            row_height = 50
            header_height = 250
            stats_height = 150
            footer_height = 80

            table_height = (len(df_dept) * row_height) + 60
            total_height = header_height + stats_height + table_height + footer_height

            bg_color = "#ffffff"
            header_bg = "#198754"
            accent_color = "#ffc107"
            light_gray = "#f8f9fa"
            padding = 40

            img = Image.new('RGB', (width, total_height), bg_color)
            draw = ImageDraw.Draw(img)

            # Fonts
            try:
                font_title = ImageFont.truetype("arialbd.ttf", 40)
                font_subtitle = ImageFont.truetype("arial.ttf", 18)
                font_big_num = ImageFont.truetype("arialbd.ttf", 55)
                font_label = ImageFont.truetype("arialbd.ttf", 16)
                font_row = ImageFont.truetype("arial.ttf", 22)
                font_row_bold = ImageFont.truetype("arialbd.ttf", 22)
                has_font = True
            except Exception:
                font_title = ImageFont.load_default()
                font_subtitle = font_title
                font_big_num = font_title
                font_label = font_title
                font_row = font_title
                font_row_bold = font_title
                has_font = False

            # Header
            draw.rectangle([(0, 0), (width, header_height)], fill=header_bg)

            # Logo
            logo_path = "logojup.png"
            if getattr(sys, 'frozen', False):
                logo_path = os.path.join(sys._MEIPASS, "logojup.png")
            if os.path.exists(logo_path):
                try:
                    logo = Image.open(logo_path).convert("RGBA")
                    logo.thumbnail((200, 100), Image.Resampling.LANCZOS)
                    img.paste(logo, (width - logo.width - padding, padding), logo)
                except Exception as e:
                    self.log(f"Erro ao carregar logo: {e}")

            draw.text((padding, padding), "Relatório Diário", font=font_title, fill="white")
            draw.text((padding, padding + 55), "de Acesso e Presença", font=font_subtitle if has_font else font_title, fill="#e8f5e9")
            draw.text((padding, header_height - 60), f"Data: {data_str}", font=font_label if has_font else font_title, fill="white")
            draw.rectangle([(padding, padding + 48), (padding + 100, padding + 51)], fill=accent_color)

            # Cards
            current_y = header_height + 20
            card_width = (width - (3 * padding)) // 2
            card_height = 120

            draw.rounded_rectangle([(padding, current_y), (padding + card_width, current_y + card_height)], radius=15, fill=light_gray, outline="#d1d1d1", width=1)
            draw.text((padding + 20, current_y + 20), "TOTAL MOVIMENTOS", font=font_label, fill="#555")
            draw.text((padding + 20, current_y + 50), str(total_mov), font=font_big_num, fill="#2c3e50")

            x_card2 = padding + card_width + padding
            draw.rounded_rectangle([(x_card2, current_y), (x_card2 + card_width, current_y + card_height)], radius=15, fill="#e8f5e9", outline="#c3e6cb", width=1)
            draw.text((x_card2 + 20, current_y + 20), "PESSOAS NA OBRA", font=font_label, fill="#155724")
            draw.text((x_card2 + 20, current_y + 50), str(total_pres), font=font_big_num, fill="#198754")

            # Tabela
            current_y += card_height + 40
            dept_col = df_dept.columns[0] if len(df_dept.columns) > 0 else 'Departamento'

            draw.text((padding, current_y), "Distribuição por Departamento", font=font_row_bold, fill="#198754")
            draw.rectangle([(padding, current_y + 35), (width - padding, current_y + 36)], fill="#eee")
            current_y += 50

            draw.text((padding + 10, current_y), "DEPARTAMENTO", font=font_label, fill="#999")
            draw.text((width - padding - 180, current_y), "QTD", font=font_label, fill="#999", anchor="ra")
            draw.text((width - padding - 50, current_y), "%", font=font_label, fill="#999", anchor="ra")
            current_y += 30

            for _, row in df_dept.iterrows():
                dept = str(row[dept_col])
                if len(dept) > 35:
                    dept = dept[:32] + "..."
                qtd = str(row['Qtd'])
                pct = f"{row['Pct']:.1f}%"

                draw.text((padding + 10, current_y + 10), dept, font=font_row, fill="#333")
                draw.text((width - padding - 180, current_y + 10), qtd, font=font_row_bold, fill="#333", anchor="ra")
                draw.text((width - padding - 50, current_y + 10), pct, font=font_row_bold, fill="#198754", anchor="ra")
                draw.line([(padding, current_y + row_height), (width - padding, current_y + row_height)], fill="#f0f0f0", width=1)
                current_y += row_height

            # Footer
            footer_y = total_height - footer_height
            draw.rectangle([(0, footer_y), (width, total_height)], fill=light_gray)
            draw.text((width // 2, footer_y + 30), "Gerado automaticamente por BotEliva", font=font_subtitle, fill="#999", anchor="mm")

            # Salvar na pasta do dia
            os.makedirs(report_dir, exist_ok=True)
            output_path = os.path.join(report_dir, f"resumo-{time_str}.png")
            img.save(output_path)
            self.log(f"Imagem resumo salva: resumo-{time_str}.png")
            return output_path

        except Exception as e:
            self.log(f"Erro ao gerar summary card: {e}")
            return None

    def _gerar_planilha_presentes(self, df_inside, report_dir, time_str):
        """
        Gera uma planilha Excel (.xlsx) formatada com a lista de pessoas
        presentes na obra, ordenadas por Departamento e Nome.

        Abas:
          - 'Presentes'  → lista completa (Departamento, Nome, CPF, Hora Entrada)
          - 'Por Depto'  → resumo de quantidade por departamento
        """
        if df_inside is None or df_inside.empty:
            self.log("Planilha Excel não gerada: nenhuma pessoa detectada dentro.")
            return None

        try:
            import openpyxl
            from openpyxl.styles import (Font, PatternFill, Alignment,
                                         Border, Side, GradientFill)
            from openpyxl.utils import get_column_letter
        except ImportError:
            self.log("openpyxl não disponível. Planilha não gerada.")
            return None

        xlsx_name = f"Presentes_Obra-{time_str}.xlsx"
        xlsx_path = os.path.join(report_dir, xlsx_name)

        now_str = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")

        # ── Preparar dados ──────────────────────────────────────────────────
        cols_wanted = {
            'Departamento': 'Departamento',
            'Nome':         'Nome Completo',
            'CPF':          'CPF',
            'Data Evento':  'Hora de Entrada / Último Acesso',
        }
        df_export = df_inside.copy()

        # Garante colunas obrigatórias
        for col in cols_wanted:
            if col not in df_export.columns:
                df_export[col] = ''

        df_export = df_export[list(cols_wanted.keys())].copy()
        df_export.columns = list(cols_wanted.values())

        # Formatar Hora de Entrada
        if 'Hora de Entrada / Último Acesso' in df_export.columns:
            df_export['Hora de Entrada / Último Acesso'] = (
                pd.to_datetime(df_export['Hora de Entrada / Último Acesso'], errors='coerce')
                .dt.strftime('%d/%m/%Y %H:%M')
            )

        # Formatar CPF (ex: 123.456.789-00) — mantém como string
        def fmt_cpf(v):
            v = str(v).strip()
            digits = ''.join(c for c in v if c.isdigit())
            if len(digits) == 11:
                return f"{digits[:3]}.{digits[3:6]}.{digits[6:9]}-{digits[9:]}"
            return v if v not in ('', 'nan', 'N/A') else '—'

        df_export['CPF'] = df_export['CPF'].apply(fmt_cpf)

        # Ordenar por Departamento e Nome
        df_export = df_export.sort_values(['Departamento', 'Nome Completo'],
                                          na_position='last').reset_index(drop=True)

        # Resumo por departamento
        df_dept_sum = (df_export.groupby('Departamento').size()
                       .reset_index(name='Qtd. Pessoas')
                       .sort_values('Qtd. Pessoas', ascending=False))

        # ── Estilos ─────────────────────────────────────────────────────────
        green_fill    = PatternFill("solid", fgColor="198754")
        lt_green_fill = PatternFill("solid", fgColor="E8F5E9")
        alt_fill      = PatternFill("solid", fgColor="F5F5F5")
        dept_fill     = PatternFill("solid", fgColor="EFF8F1")
        header_font   = Font(bold=True, color="FFFFFF", size=11)
        dept_font     = Font(bold=True, color="155724", size=11)
        bold_font     = Font(bold=True, size=10)
        normal_font   = Font(size=10)
        center        = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left          = Alignment(horizontal='left', vertical='center', wrap_text=True)

        thin = Side(border_style="thin", color="D0D0D0")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def apply_header(ws, cols, row_index=1):
            ws.row_dimensions[row_index].height = 32
            for ci, col_name in enumerate(cols, 1):
                cell = ws.cell(row=row_index, column=ci, value=col_name)
                cell.font = header_font
                cell.fill = green_fill
                cell.alignment = center
                cell.border = border

        def auto_width(ws, col_index, min_w=12, max_w=50):
            max_len = 0
            for row in ws.iter_rows(min_col=col_index, max_col=col_index):
                for cell in row:
                    try:
                        max_len = max(max_len, len(str(cell.value) or ''))
                    except Exception:
                        pass
            ws.column_dimensions[get_column_letter(col_index)].width = min(max(max_len + 3, min_w), max_w)

        # ── Criar workbook ────────────────────────────────────────────────
        wb = openpyxl.Workbook()

        # ── Aba 1: Presentes ─────────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Presentes"

        # Linha de título
        ws1.merge_cells("A1:D1")
        title_cell = ws1["A1"]
        title_cell.value = f"RELATÓRIO DE PRESENÇA — {now_str}"
        title_cell.font = Font(bold=True, color="FFFFFF", size=13)
        title_cell.fill = PatternFill("solid", fgColor="06402B")
        title_cell.alignment = center
        ws1.row_dimensions[1].height = 35

        # Linha de totais
        ws1.merge_cells("A2:D2")
        tot_cell = ws1["A2"]
        tot_cell.value = f"Total de pessoas na obra: {len(df_export)}"
        tot_cell.font = Font(bold=True, color="155724", size=11)
        tot_cell.fill = lt_green_fill
        tot_cell.alignment = center
        ws1.row_dimensions[2].height = 24

        # Linha em branco
        ws1.row_dimensions[3].height = 5

        # Cabeçalho (linha 4)
        headers = list(df_export.columns)
        apply_header(ws1, headers, row_index=4)

        # Dados — agrupados com linha separadora por departamento
        row_idx = 5
        current_dept = None
        fill_toggle = False

        for _, data_row in df_export.iterrows():
            dept = data_row['Departamento']

            # Separador de departamento
            if dept != current_dept:
                current_dept = dept
                fill_toggle = False
                # Linha de label do departamento
                ws1.merge_cells(f"A{row_idx}:D{row_idx}")
                dept_cell = ws1[f"A{row_idx}"]
                dept_cell.value = f"  {dept}".upper()
                dept_cell.font = dept_font
                dept_cell.fill = dept_fill
                dept_cell.alignment = left
                dept_cell.border = border
                ws1.row_dimensions[row_idx].height = 22
                row_idx += 1

            # Linha de dados
            fill = alt_fill if fill_toggle else PatternFill("solid", fgColor="FFFFFF")
            fill_toggle = not fill_toggle

            values = [dept, data_row['Nome Completo'],
                      data_row['CPF'],
                      data_row['Hora de Entrada / Último Acesso']]

            for ci, val in enumerate(values, 1):
                cell = ws1.cell(row=row_idx, column=ci, value=val)
                cell.font = normal_font
                cell.fill = fill
                cell.alignment = left if ci == 2 else center
                cell.border = border
            ws1.row_dimensions[row_idx].height = 20
            row_idx += 1

        # Ajustar larguras
        for ci in range(1, len(headers) + 1):
            auto_width(ws1, ci)
        ws1.freeze_panes = "A5"  # congela cabeçalho

        # ── Aba 2: Resumo por Departamento ──────────────────────────────
        ws2 = wb.create_sheet("Por Depto")

        ws2.merge_cells("A1:B1")
        ws2["A1"].value = "RESUMO POR DEPARTAMENTO"
        ws2["A1"].font = Font(bold=True, color="FFFFFF", size=13)
        ws2["A1"].fill = PatternFill("solid", fgColor="06402B")
        ws2["A1"].alignment = center
        ws2.row_dimensions[1].height = 32

        apply_header(ws2, ["Departamento", "Qtd. Pessoas"], row_index=2)

        for ri, (_, dr) in enumerate(df_dept_sum.iterrows(), 3):
            fill = alt_fill if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            ws2.cell(row=ri, column=1, value=dr['Departamento']).font = bold_font
            ws2.cell(row=ri, column=1).fill = fill
            ws2.cell(row=ri, column=1).border = border
            ws2.cell(row=ri, column=1).alignment = left

            ws2.cell(row=ri, column=2, value=int(dr['Qtd. Pessoas'])).font = bold_font
            ws2.cell(row=ri, column=2).fill = fill
            ws2.cell(row=ri, column=2).border = border
            ws2.cell(row=ri, column=2).alignment = center
            ws2.row_dimensions[ri].height = 20

        auto_width(ws2, 1, min_w=25)
        auto_width(ws2, 2, min_w=15)

        # ── Salvar ───────────────────────────────────────────────────────
        wb.save(xlsx_path)
        self.log(f"✅ Planilha Excel salva: {xlsx_name} ({len(df_export)} pessoas)")
        return xlsx_path

    def send_email(self, report_data):

        self.status("Enviando Email...")

        if isinstance(report_data, str):
            return f"Erro nos Dados: {report_data}"

        sender_email = self.config.get("smtp_user")
        recipients_raw = self.config.get("email_recipients", "")
        if not recipients_raw:
            return "Erro: Nenhum destinatário configurado."
        recipients_list = [e.strip() for e in recipients_raw.split(",") if e.strip()]

        password = self.config.get("smtp_password")
        smtp_server = self.config.get("smtp_server")
        smtp_port = self.config.get("smtp_port")

        if not all([sender_email, recipients_list, password, smtp_server, smtp_port]):
            return "Email não configurado completamente."

        try:
            server = smtplib.SMTP(smtp_server, int(smtp_port))
            server.starttls()
            server.login(sender_email, password)

            count_sent = 0
            for receiver_email in recipients_list:
                msg = MIMEMultipart('related')
                msg['From'] = sender_email
                msg['To'] = receiver_email
                msg['Subject'] = f"📊 Relatório Diário de Obra — {datetime.datetime.now().strftime('%d/%m/%Y')}"

                msg_alternative = MIMEMultipart('alternative')
                msg.attach(msg_alternative)

                msg_text = MIMEText(report_data['html'], 'html')
                msg_alternative.attach(msg_text)

                # Imagens inline (gráficos)
                for c_id, img_data in report_data['images'].items():
                    img_part = MIMEBase('image', 'png')
                    img_part.set_payload(img_data)
                    encoders.encode_base64(img_part)
                    img_part.add_header('Content-ID', f'<{c_id}>')
                    # FIX: removida a linha duplicada de Content-Disposition
                    img_part.add_header('Content-Disposition', 'inline', filename=f'{c_id}.png')
                    msg.attach(img_part)

                # Imagem resumo como anexo
                summary_path = report_data.get('summary_image')
                if summary_path and os.path.exists(summary_path):
                    try:
                        with open(summary_path, 'rb') as f:
                            part = MIMEBase('image', 'png')
                            part.set_payload(f.read())
                            encoders.encode_base64(part)
                            part.add_header('Content-Disposition', 'attachment; filename="Resumo_Obra.png"')
                            msg.attach(part)
                            self.log("Imagem resumo anexada.")
                    except Exception as e:
                        self.log(f"Erro ao anexar resumo: {e}")

                # Planilha Excel com presentes (por departamento)
                xlsx_path = report_data.get('xlsx_presentes')
                if xlsx_path and os.path.exists(xlsx_path):
                    try:
                        with open(xlsx_path, 'rb') as f:
                            part_xlsx = MIMEBase('application',
                                'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                            part_xlsx.set_payload(f.read())
                            encoders.encode_base64(part_xlsx)
                            fname = os.path.basename(xlsx_path)
                            part_xlsx.add_header('Content-Disposition',
                                f'attachment; filename="{fname}"')
                            msg.attach(part_xlsx)
                            self.log(f"Planilha Excel anexada: {fname}")
                    except Exception as e:
                        self.log(f"Erro ao anexar planilha Excel: {e}")

                server.sendmail(sender_email, receiver_email, msg.as_string())
                count_sent += 1

            server.quit()
            return f"✅ Relatório enviado para {count_sent} destinatário(s)!"

        except Exception as e:
            import traceback
            traceback.print_exc()
            return f"Erro envio email: {e}"

    def stop(self):
        self.stop_flag = True
        self.log("Parando serviço...")

    def run(self):
        self.log("🚀 Serviço de Monitoramento Iniciado.")
        schedule.clear()

        sched_type = self.config.get("schedule_type", "Diário").lower()
        times_str = self.config.get("schedule_times", "08:00")

        def pad_time(t_str):
            t_str = t_str.strip()
            if not t_str:
                return "08:00"
            if ":" in t_str:
                h, m = t_str.split(":", 1)
                try:
                    return f"{int(h):02d}:{int(m):02d}"
                except:
                    return "08:00"
            else:
                try:
                    return f"{int(t_str):02d}:00"
                except:
                    return "08:00"

        raw_times = [t for t in times_str.split(",") if t.strip()]
        if not raw_times:
            raw_times = ["08:00"]
        times_list = [pad_time(t) for t in raw_times]

        days_str = self.config.get("schedule_days", "")
        days_list = [d.strip().lower() for d in days_str.split(",") if d.strip()]

        for t in times_list:
            if "diario" in sched_type or "diário" in sched_type:
                schedule.every().day.at(t).do(self.run_task)
                self.log(f"⏰ Agendado: Diariamente às {t}")

            elif "semanal" in sched_type:
                day_map = {
                    "seg": schedule.every().monday,
                    "ter": schedule.every().tuesday,
                    "qua": schedule.every().wednesday,
                    "qui": schedule.every().thursday,
                    "sex": schedule.every().friday,
                    "sab": schedule.every().saturday,
                    "sáb": schedule.every().saturday,
                    "dom": schedule.every().sunday,
                }
                if not days_list:
                    schedule.every().monday.at(t).do(self.run_task)
                    self.log(f"⏰ Agendado: Semanalmente (Seg) às {t}")
                else:
                    for d in days_list:
                        for k, v in day_map.items():
                            if k in d:
                                v.at(t).do(self.run_task)
                                self.log(f"⏰ Agendado: Semanalmente ({k.capitalize()}) às {t}")
                                break

            elif "mensal" in sched_type:
                schedule.every().day.at(t).do(self.run_task_monthly, days_list)
                self.log(f"⏰ Agendado: Mensalmente nos dias '{days_str}' às {t}")

        def _update_next_run_ui():
            nxt = schedule.next_run()
            if nxt and self.update_stats_callback:
                self.update_stats_callback(self.success_count, self.fail_count, nxt.strftime('%d/%m %H:%M'))

        _update_next_run_ui()
        self.status("⏳ Aguardando próximo horário...")

        while not self.stop_flag:
            schedule.run_pending()
            _update_next_run_ui()
            time.sleep(1)

        self.status("⛔ Serviço Parado.")

    def run_task_monthly(self, days_list):
        hoje = str(datetime.datetime.now().day)
        if not days_list:
            if hoje == "1":
                self.run_task()
        else:
            if hoje in days_list:
                self.run_task()

    def run_task(self):
        if self.stop_flag:
            return

        if self.clear_log_callback:
            self.clear_log_callback()

        if self.update_stats_callback:
            self.update_stats_callback(self.success_count, self.fail_count, "Executando...")

        self.log(f"--- ⚡ Iniciando Ciclo: {datetime.datetime.now().strftime('%d/%m %H:%M')} ---")

        # login_and_export_report agora devolve dict ou None
        paths = self.login_and_export_report()
        if not paths or self.stop_flag:
            if not self.stop_flag:
                self.fail_count += 1
                self.log("❌ Falha na extração. Aguardando próximo ciclo.")
            self.status("Aguardando...")
            return

        report = self.process_data(paths)
        if isinstance(report, str):
            self.fail_count += 1
            self.log(f"❌ {report}")
            self.status("Aguardando...")
            return

        self.log("📊 Relatório gerado. Enviando email...")

        email_status = self.send_email(report)
        if "Erro" in str(email_status):
            self.fail_count += 1
        else:
            self.success_count += 1

        self.log(email_status)
        self.log("✅ Ciclo finalizado. Aguardando próximo horário.")
        self.status("Aguardando...")
